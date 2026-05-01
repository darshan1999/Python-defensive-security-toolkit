#!/usr/bin/env python3
"""
Multi-Sheet XLSX Exporter - Export JSON to Excel with multiple sheets.

Accepts JSON with categories (network, endpoint, email, cloud).
Creates sheet per category + Summary sheet.
Uses openpyxl for formatting (bold headers, severity color-coding).
"""

import argparse
import json

def export_to_xlsx(json_file, output_file):
    """Export JSON data to multi-sheet Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("[!] openpyxl not installed. Use: pip install openpyxl", flush=True)
        print("[*] Creating JSON output instead...", flush=True)
        with open(output_file.replace('.xlsx', '.json'), 'w') as f:
            with open(json_file, 'r') as src:
                f.write(src.read())
        return
    
    # Load JSON data
    try:
        with open(json_file, 'r') as f:
            data = json.load(f)
    except Exception as e:
        print(f"[!] Error loading JSON: {e}", flush=True)
        return
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Color fills by severity
    severity_colors = {
        'CRITICAL': 'FF0000',
        'HIGH': 'FFA500',
        'MEDIUM': 'FFFF00',
        'LOW': '00FF00',
        'INFO': '0099FF'
    }
    
    # Create Summary sheet
    ws_summary = wb.create_sheet("Summary", 0)
    ws_summary['A1'] = "Security Report Summary"
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary['A2'] = f"Generated: {json.get('timestamp', 'Unknown')}"
    ws_summary['A3'] = f"Total Records: {len(data.get('records', data.get('alerts', [])))}"
    
    # Categorize data
    categories = {}
    
    if isinstance(data, dict):
        # Check for category structure
        if 'network' in data or 'endpoint' in data or 'email' in data or 'cloud' in data:
            categories = {k: v for k, v in data.items() if isinstance(v, list)}
        elif 'alerts' in data or 'records' in data or 'items' in data:
            # Find the list of items
            items_key = next((k for k in data.keys() if isinstance(data[k], list)), None)
            if items_key:
                categories['data'] = data[items_key]
        else:
            categories['data'] = [data]
    
    # Create sheets per category
    for category, items in categories.items():
        if not items or not isinstance(items, list):
            continue
        
        ws = wb.create_sheet(category.capitalize())
        
        # Get column headers from first item
        if items and isinstance(items[0], dict):
            headers = list(items[0].keys())
            
            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Write data
            for row, item in enumerate(items, 2):
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col)
                    value = item.get(header, '')
                    cell.value = value
                    
                    # Color code by severity
                    if header.lower() == 'severity' and value in severity_colors:
                        cell.fill = PatternFill(start_color=severity_colors[value], 
                                               end_color=severity_colors[value], 
                                               fill_type="solid")
            
            # Auto-adjust column widths
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[chr(64 + col)].width = 15
    
    # Save workbook
    try:
        wb.save(output_file)
        print(f"[+] Excel file saved to {output_file}")
    except Exception as e:
        print(f"[!] Error saving Excel: {e}", flush=True)

def main():
    parser = argparse.ArgumentParser(
        description='Export JSON to multi-sheet Excel with formatting.',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python3 multi_sheet_xlsx_exporter.py --input alerts.json --output report.xlsx
  python3 multi_sheet_xlsx_exporter.py --input data.json --output security_report.xlsx
        """)
    
    parser.add_argument('--input', type=str, required=True, help='Input JSON file')
    parser.add_argument('--output', type=str, required=True, help='Output XLSX file')
    
    args = parser.parse_args()
    
    print(f"[*] Exporting {args.input} to {args.output}...", flush=True)
    export_to_xlsx(args.input, args.output)

if __name__ == "__main__":
    main()
