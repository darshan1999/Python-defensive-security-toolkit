"""
Exporting Data To Dashboards Csv Excel Json

Security operations automation tool.
"""

import csv
# Simulated normalized logs
events = [
    {"timestamp": "2025-06-14T16:22:00", "source": "auth", "event": "failed_login", "user": "alice", "ip": "10.0.0.5"},
    {"timestamp": "2025-06-14T16:22:10", "source": "web", "event": "url_access", "user": "alice", "url": "/admin", "ip": "10.0.0.5"}
]
with open("events_export.csv", "w", newline="") as csvfile:
    fieldnames = events[0].keys()
    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
    writer.writeheader()  # Write column names
    for event in events:
        writer.writerow(event)
# - `DictWriter` handles writing dictionary keys as column headers.
# - Ensure all dictionaries in the list have the same keys to avoid errors.
#
# The Excel format enables richer formatting, conditional highlights, charts, and
# multiple sheets. Python provides `openpyxl` and `xlsxwriter` for this.
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
wb = Workbook()
ws = wb.active
ws.title = "Log Events"
# Header row
headers = list(events[0].keys())
ws.append(headers)
# Bold + fill style
header_style = Font(bold=True)
for col in range(1, len(headers)+1):
    ws.cell(row=1, column=col).font = header_style
    ws.cell(row=1, column=col).fill = PatternFill(start_color="D9E1F2", fill_type="solid")
# Write rows
for event in events:
    ws.append([event.get(h, "") for h in headers])
wb.save("events_export.xlsx")
# - Use conditional formatting to highlight suspicious rows (e.g., failed logins, access
#   to `/admin`)
# - Each worksheet can represent a log source (e.g., Auth, Web, EDR)
#
# JSON is ideal for nested, structured data and interoperability between systems.
import json
with open("events_export.json", "w") as f:
    json.dump(events, f, indent=2)
# - Use `indent=2` for human-readable formatting.
# - JSON supports nested fields (like lists, dictionaries inside logs).
#
# Suppose we have logs from different sources and want to split them into sheets.
from openpyxl import Workbook
sources = {
    "auth": [e for e in events if e["source"] == "auth"],
    "web": [e for e in events if e["source"] == "web"]
}
wb = Workbook()
for i, (source, rows) in enumerate(sources.items()):
    ws = wb.create_sheet(title=source) if i > 0 else wb.active
    ws.title = source
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
wb.save("multisource_logs.xlsx")
# Dashboards often ingest data via:
# - REST APIs: Feed JSON or CSV data into tools like Splunk HEC, Elastic Beats, Graylog
#   Input API
# - File Watchers: Export files to watched directories
# - Database Sync: Store to SQLite/Postgres and connect visualization tools (e.g.,
#   Metabase, Grafana)
#
#
#
# After correlation analysis
alerts = [
    {"ip": "10.0.0.5", "issue": "failed login + /admin access", "user": "alice", "time": "2025-06-14T16:22:10"},
    {"ip": "10.0.0.8", "issue": "multiple IPs for one user", "user": "bob", "time": "2025-06-14T18:12:50"}
]
# Export as CSV
with open("correlation_alerts.csv", "w", newline="") as f:
    writer = csv.DictWriter(f, fieldnames=alerts[0].keys())
    writer.writeheader()
    writer.writerows(alerts)
# Mastering data export techniques:
# - Improves team visibility and response time
# - Enables collaboration between SOC, IR, and compliance teams
# - Supports integration with tools across the detection and response ecosystem