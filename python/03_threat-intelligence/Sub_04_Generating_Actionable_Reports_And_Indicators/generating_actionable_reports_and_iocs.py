"""
Generating Actionable Reports And Iocs

Security operations automation tool.
"""

def build_report(ioc_data):
    return {
        "ioc": ioc_data.get("ioc"),
        "type": ioc_data.get("type"),
        "source": ioc_data.get("sources", []),
        "reputation": ioc_data.get("abuse_score", 0),
        "classification": classify_threat(ioc_data),
        "geolocation": {
            "country": ioc_data.get("country"),
            "asn": ioc_data.get("asn")
        },
        "vt_malicious": ioc_data.get("virustotal_malicious", 0),
        "related_entities": ioc_data.get("related", []),
        "mitre_ttps": ioc_data.get("mitre", []),
        "recommendations": generate_response(ioc_data)
    }
def classify_threat(ioc):
    score = ioc.get("abuse_score", 0)
    if score >= 85 or ioc.get("vt_malicious", 0) > 10:
        return "High"
    elif score >= 50:
        return "Medium"
    else:
        return "Low"
def generate_response(ioc):
    if classify_threat(ioc) == "High":
        return ["Block IOC at perimeter", "Isolate affected asset", "Create SIEM rule"]
    elif classify_threat(ioc) == "Medium":
        return ["Monitor activity", "Flag IOC in threat feeds"]
    else:
        return ["No action required", "Log for future correlation"]
# ## 4. Exporting Reports in JSON and CSV
import json
report = build_report(enriched_ioc)
with open("ioc_report.json", "w") as f:
    json.dump(report, f, indent=2)
import csv
with open("ioc_report.csv", "w", newline='') as f:
    writer = csv.DictWriter(f, fieldnames=report.keys())
    writer.writeheader()
    writer.writerow(report)
# ## 5. Exporting to Excel using `openpyxl`
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "IOC Report"
for i, (k, v) in enumerate(report.items(), start=1):
    ws[f"A{i}"] = k
    ws[f"B{i}"] = str(v)
wb.save("ioc_report.xlsx")
# Excel is useful for sharing reports with non-technical stakeholders.
#
# ## 6. Exporting to Markdown and HTML
md = f"""# IOC Report: {report['ioc']}
**Type:** {report['type']}  
**Reputation Score:** {report['reputation']}  
**Country:** {report['geolocation']['country']}  
**ASN:** {report['geolocation']['asn']}  
**VirusTotal Detections:** {report['vt_malicious']}  
**Classification:** {report['classification']}
## Recommendations
"""
for rec in report["recommendations"]:
    md += f"- {rec}\n"
with open("ioc_report.md", "w") as f:
    f.write(md)
# Convert Markdown to HTML using `markdown` library:
import markdown
html = markdown.markdown(md)
with open("ioc_report.html", "w") as f:
    f.write(html)
# ## 7. Exporting to PDF
# Use `reportlab` or convert HTML to PDF with `pdfkit` and `wkhtmltopdf`.
import pdfkit
pdfkit.from_file("ioc_report.html", "ioc_report.pdf")
# ## 8. Grouping and Aggregating Multiple IOC Reports
all_reports = []
for ioc in ioc_list:
    enriched = enrich_ip(ioc)
    rep = build_report(enriched)
    all_reports.append(rep)
with open("all_ioc_reports.json", "w") as f:
    json.dump(all_reports, f, indent=2)
# You can group IOCs by severity or type before exporting.
#
# ## 9. Including MITRE ATT&CK Tags
# Add TTPs based on API detection (e.g., phishing -> T1566, command and control ->
# T1071):
def tag_mitre(ioc):
    if "phishing" in ioc.get("related_tags", []):
        return ["T1566"]
    if "c2" in ioc.get("behavior", ""):
        return ["T1071"]
    return []
# This helps align with MITRE frameworks in SIEMs or reports.
#
# ## 10. Delivering Actionable Intelligence
# Ensure that each report answers these questions:
# - What is the IOC?
# - How dangerous is it?
# - What is it associated with?
# - What should we do about it?
# Make your report readable by both humans and machines.
#
# ## 11. Sending Reports via Email (Optional)
import smtplib
from email.message import EmailMessage
msg = EmailMessage()
msg["Subject"] = "IOC Alert Report"
msg["From"] = "analyst@example.com"
msg["To"] = "soc@example.com"
msg.set_content(md)
with open("ioc_report.xlsx", "rb") as f:
    msg.add_attachment(f.read(), maintype="application", subtype="xlsx", filename="ioc_report.xlsx")
s = smtplib.SMTP("smtp.example.com")
s.send_message(msg)
s.quit()